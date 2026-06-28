import os
import requests
import pandas as pd
import openpyxl
from flask import Flask, request, jsonify
from difflib import get_close_matches

app = Flash(__name__)

VERIFY_TOKEN = os.environ.get("WA_VERIFY_TOKEN", "bodega_inventario_2024")
ACCESS_TOKEN = os.environ.get("WA_ACCESS_TOKEN", "")
PHONE_ID     = os.environ.get("WA_PHONE_NUMBER_ID", "")
EXCEL_PATH   = os.environ.get("EXCEL_PATH", "Inventario_Bodega.xlsx")

# Columnas reales del Excel (con tildes)
COL_NOMBRE  = "Nombre del Ãtem"
COL_CAT     = "CategorÃ­a"
COL_ESTADO  = "Estado"
COL_TOTAL   = "Stock Total"
COL_DISP    = "Disponible"
COL_USO     = "En Uso"
COL_REP     = "En ReparaciÃ³n"
COL_UBIC    = "UbicaciÃ³n"

def cargar_inventario():
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Inventario", header=1)
        df.columns = df.columns.str.strip()
        return df.dropna(subset=["ID"])
    except Exception:
        return None

def buscar_item(df, consulta):
    consulta = consulta.lower().strip()
    match_id = df[df["ID"].str.lower() == consulta]
    if not match_id.empty:
        return match_id.iloc[0]
    nombres = df[COL_NOMBRE].str.lower().tolist()
    ids = df["ID"].str.lower().tolist()
    cercanos = get_close_matches(consulta, nombres + ids, n=1, cutoff=0.4)
    if cercanos:
        match = df[
            (df[COL_NOMBRE].str.lower() == cercanos[0]) |
            (df["ID"].str.lower() == cercanos[0])
        ]
        if not match.empty:
            return match.iloc[0]
    match_cont = df[df[COL_NOMBRE].str.lower().str.contains(consulta, na=False)]
    if not match_cont.empty:
        return match_cont.iloc[0]
    return None

def formato_item(item):
    estado = str(item.get(COL_ESTADO, ""))
    emoji = {"OK": "â", "En uso": "ðµ", "Parcial": "ð¡", "CrÃ­tico": "ð´"}.get(estado, "âª")
    disp = item.get(COL_DISP, 0)
    try:
        disp = int(float(disp)) if pd.notna(disp) else 0
    except Exception:
        disp = 0
    return (
        f"{emoji} *{item[COL_NOMBRE]}* ({item['ID']})\n"
        f"â¢ CategorÃ­a: {item.get(COL_CAT, '')}\n"
        f"â¢ Stock total: {int(item.get(COL_TOTAL, 0))}\n"
        f"â¢ Disponible: {disp}\n"
        f"â¢ En uso: {int(item.get(COL_USO, 0))}\n"
        f"â¢ En reparaciÃ³n: {int(item.get(COL_REP, 0))}\n"
        f"â¢ Estado: {estado}\n"
        f"â¢ UbicaciÃ³n: {item.get(COL_UBIC, '')}"
    )

AYUDA_AGREGAR = (
    "ð *Formato para agregar un Ã­tem:*\n\n"
    "agregar ID | Nombre | CategorÃ­a | Stock Total | Disponible | En Uso | En ReparaciÃ³n | Estado | UbicaciÃ³n\n\n"
    "ð *Ejemplo:*\n"
    "_agregar H012 | Martillo grande | Herramientas | 5 | 4 | 1 | 0 | OK | Estante A_\n\n"
    "ð *CategorÃ­as vÃ¡lidas:* Herramientas, Insumos, EPP, Repuestos\n"
    "ð *Estados vÃ¡lidos:* OK, En uso, Parcial, CrÃ­tico"
)

def agregar_item(texto):
    """
    Formato obligatorio (9 campos separados por |):
    ID | Nombre | CategorÃ­a | Stock Total | Disponible | En Uso | En ReparaciÃ³n | Estado | UbicaciÃ³n
    """
    partes = [p.strip() for p in texto.split("|")]

    if len(partes) < 9:
        return (
            f"â Faltan campos. Se necesitan exactamente 9 campos.\n\n"
            + AYUDA_AGREGAR
        )

    item_id    = partes[0]
    nombre     = partes[1]
    categoria  = partes[2]
    stock_str  = partes[3]
    disp_str   = partes[4]
    uso_str    = partes[5]
    rep_str    = partes[6]
    estado     = partes[7]
    ubicacion  = partes[8]

    # Validar campos vacÃ­os
    campos_vacios = [
        ("ID", item_id), ("Nombre", nombre), ("CategorÃ­a", categoria),
        ("Stock Total", stock_str), ("Disponible", disp_str),
        ("En Uso", uso_str), ("En ReparaciÃ³n", rep_str),
        ("Estado", estado), ("UbicaciÃ³n", ubicacion)
    ]
    for campo, valor in campos_vacios:
        if not valor:
            return f"â El campo *{campo}* no puede estar vacÃ­o.\n\n" + AYUDA_AGREGAR

    # Validar numÃ©ricos
    try:
        stock = int(stock_str)
        disp  = int(disp_str)
        uso   = int(uso_str)
        rep   = int(rep_str)
    except ValueError as e:
        return f"â Stock Total, Disponible, En Uso y En ReparaciÃ³n deben ser nÃºmeros enteros.\n\nError: {e}"

    # Validar categorÃ­a
    cats_validas = ["Herramientas", "Insumos", "EPP", "Repuestos"]
    if categoria not in cats_validas:
        return f"â CategorÃ­a '{categoria}' no vÃ¡lida.\nUsa: {', '.join(cats_validas)}"

    # Validar estado
    estados_validos = ["OK", "En uso", "Parcial", "CrÃ­tico"]
    if estado not in estados_validos:
        return f"â Estado '{estado}' no vÃ¡lido.\nUsa: {', '.join(estados_validos)}"

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Inventario"]
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=item_id)
        ws.cell(row=next_row, column=2, value=nombre)
        ws.cell(row=next_row, column=3, value=categoria)
        ws.cell(row=next_row, column=4, value=stock)
        ws.cell(row=next_row, column=5, value=disp)
        ws.cell(row=next_row, column=6, value=uso)
        ws.cell(row=next_row, column=7, value=rep)
        ws.cell(row=next_row, column=8, value=estado)
        ws.cell(row=next_row, column=9, value=ubicacion)
        wb.save(EXCEL_PATH)
        return (
            f"â *Ãtem agregado al inventario:*\n"
            f"â¢ ID: {item_id}\n"
            f"â¢ Nombre: {nombre}\n"
            f"â¢ CategorÃ­a: {categoria}\n"
            f"â¢ Stock Total: {stock}\n"
            f"â¢ Disponible: {disp}\n"
            f"â¢ En Uso: {uso}\n"
            f"â¢ En ReparaciÃ³n: {rep}\n"
            f"â¢ Estado: {estado}\n"
            f"â¢ UbicaciÃ³n: {ubicacion}"
        )
    except Exception as e:
        return f"â Error al guardar en el inventario: {str(e)}"

def responder(texto):
    t = texto.lower().strip()

    if t.startswith("agregar "):
        contenido = texto[8:].strip()
        return agregar_item(contenido)

    if t == "agregar":
        return AYUDA_AGREGAR

    df = cargar_inventario()
    if df is None:
        return "â No pude leer el inventario. Verifica el archivo Excel."

    if t in ["hola", "hi", "ayuda", "help", "menu", "inicio", "menÃº"]:
        return (
            "ð Hola! Soy *Maquina*, tu bot de inventario de bodega.\n\n"
            "Puedes preguntarme:\n"
            "â¢ *stock [Ã­tem]* â ver disponibilidad\n"
            "â¢ *buscar [Ã­tem]* â buscar por nombre o ID\n"
            "â¢ *agregar* â instrucciones para agregar Ã­tem\n"
            "â¢ *reparaciÃ³n* â Ã­tems en reparaciÃ³n\n"
            "â¢ *en uso* â Ã­tems en uso\n"
            "â¢ *crÃ­tico* â Ã­tems con stock crÃ­tico\n"
            "â¢ *resumen* â dashboard general\n"
            "â¢ *herramientas / insumos / epp / repuestos*\n\n"
            "Ejemplo: _stock taladro_ o _buscar H001_"
        )

    if t in ["resumen", "total", "dashboard"]:
        total = len(df)
        ok  = (df[COL_ESTADO] == "OK").sum()
        uso = (df[COL_ESTADO] == "En uso").sum()
        rep = df[COL_ESTADO].isin(["Parcial", "CrÃ­tico"]).sum()
        cats = df[COL_CAT].value_counts()
        lineas = [f"ð *Resumen de inventario*\nTotal: {total} | â OK: {ok} | ðµ En uso: {uso} | ð¡ Incidencias: {rep}\n"]
        for cat, n in cats.items():
            lineas.append(f"  â¢ {cat}: {n} Ã­tems")
        return "\n".join(lineas)

    if t in ["reparacion", "reparaciÃ³n", "en reparacion", "en reparaciÃ³n", "reparar"]:
        sub = df[df[COL_REP] > 0]
        if sub.empty:
            return "â No hay Ã­tems en reparaciÃ³n actualmente."
        lineas = [f"ð§ *Ãtems en reparaciÃ³n ({len(sub)}):*"]
        for _, row in sub.iterrows():
            lineas.append(f"  â¢ {row[COL_NOMBRE]} ({row['ID']}): {int(row[COL_REP])} unid.")
        return "\n".join(lineas)

    if t in ["en uso", "uso"]:
        sub = df[df[COL_USO] > 0]
        if sub.empty:
            return "â¹ï¸ No hay Ã­tems actualmente en uso."
        lineas = [f"ðµ *Ãtems en uso ({len(sub)}):*"]
        for _, row in sub.iterrows():
            lineas.append(f"  â¢ {row[COL_NOMBRE]} ({row['ID']}): {int(row[COL_USO])} unid.")
        return "\n".join(lineas)

    if t in ["critico", "crÃ­tico", "stock critico", "stock crÃ­tico"]:
        sub = df[df[COL_ESTADO].isin(["CrÃ­tico"])]
        if sub.empty:
            return "â No hay Ã­tems en estado crÃ­tico."
        lineas = [f"ð´ *Ãtems crÃ­ticos ({len(sub)}):*"]
        for _, row in sub.iterrows():
            lineas.append(f"  â¢ {row[COL_NOMBRE]} ({row['ID']}): stock {int(row[COL_TOTAL])}")
        return "\n".join(lineas)

    for key, cat in {"herramientas": "Herramientas", "insumos": "Insumos", "epp": "EPP", "repuestos": "Repuestos"}.items():
        if t == key:
            sub = df[df[COL_CAT] == cat]
            if sub.empty:
                return f"No hay Ã­tems en {cat}."
            lineas = [f"ð¦ *{cat} ({len(sub)} Ã­tems):*"]
            for _, row in sub.iterrows():
                disp = row.get(COL_DISP, 0)
                try:
                    disp = int(float(disp)) if pd.notna(disp) else 0
                except Exception:
                    disp = 0
                lineas.append(f"  â¢ {row[COL_NOMBRE]} ({row['ID']}): {disp} disp.")
            return "\n".join(lineas)

    for prefix in ["stock ", "buscar ", "ver ", "info "]:
        if t.startswith(prefix):
            consulta = t[len(prefix):]
            item = buscar_item(df, consulta)
            if item is not None:
                return formato_item(item)
            return f"ð No encontrÃ© Ã­tem con '{consulta}'.\nIntenta con el ID (ej: H001) o parte del nombre."

    item = buscar_item(df, t)
    if item is not None:
        return formato_item(item)

    return "ð¤ No entendÃ­ esa consulta.\nEscribe *ayuda* para ver los comandos disponibles."

def send_whatsapp_message(to, body):
    url = f"https://graph.facebook.com/v19.0/{PHONE_ID}/messages"
    r = requests.post(
        url,
        headers={"Authorization": f"Bearer {ACCESS_TOKEN}", "Content-Type": "application/json"},
        json={"messaging_product": "whatsapp", "to": to, "type": "text", "text": {"body": body}},
        timeout=10
    )
    return r.status_code, r.json()

@app.route("/webhook", methods=["GET"])
def verify():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    if mode == "subscribe" and token == VERIFY_TOKEN:
        return challenge, 200
    return "Forbidden", 403

@app.route("/webhook", methods=["POST"])
def webhook():
    data = request.get_json()
    try:
        msg = data["entry"][0]["changes"][0]["value"]["messages"][0]
        body = msg.get("text", {}).get("body", "")
        if body:
            respuesta = responder(body)
            send_whatsapp_message(msg["from"], respuesta)
    except (KeyError, IndexError):
        pass
    return '{"status":"ok"}', 200

@app.route("/", methods=["GET"])
def index():
    return "Maquina - Bot de inventario activo â", 200

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
